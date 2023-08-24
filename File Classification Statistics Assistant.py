import os
from termcolor import colored
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def categorize_files_to_excel(folder_path, excel_filename):
    # 創建一個新的 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "File Classification"  # 設置工作表的標題為 "File Classification"

    header_font = Font(bold=True, size=14, color="000000")
    # 目录行背景颜色设置（这里使用绿色作为示例）
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # 遍历资料夹中的目录和檔案
    for root, _, files in os.walk(folder_path):
        parent_folder = os.path.basename(root)  # 获取父目录名称
        file_count = len(files)  # 计算文件数量

        # 将目录名称和文件数量写入工作表中，同时设置背景颜色
        ws.append([f"Folder: {parent_folder}                        Total Files: {file_count}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=1)
        ws.cell(row=ws.max_row, column=1).fill = green_fill  # 设置目录行背景颜色
        ws.cell(row=ws.max_row, column=1).font = header_font
        ws.cell(row=ws.max_row, column=2).font = header_font
        ws.cell(row=ws.max_row, column=3).font = header_font

        # 将每个文件名称写入工作表中
        for filename in files:
            ws.append([filename])

    # 調整列寬，確保內容能夠適應儲存格
    for row in ws.iter_rows():
        for cell in row:
            if cell.coordinate in ws.merged_cells:
                continue
            try:
                column_width = len(str(cell.value)) + 20  # 考虑额外的宽度
                if column_width > ws.column_dimensions[cell.column_letter].width:
                    ws.column_dimensions[cell.column_letter].width = column_width
            except:
                pass

    # 根據當前時間生成更新後的 Excel 檔案名稱
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    updated_excel_filename = f"{excel_filename[:-5]}_{current_time}.xlsx"
    wb.save(updated_excel_filename)

if __name__ == "__main__":
    path_to_A_folder = r"C:\Users\DQE\Desktop\testautobot\findaida64\找出來的圖\20230707"
    excel_output_filename = "file_classification.xlsx"
    categorize_files_to_excel(path_to_A_folder, excel_output_filename)
